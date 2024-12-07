from openpyxl import Workbook, load_workbook
import os


def save_to_xlsx(data, filename):
    '''
    Save data to an XLSX file.
    '''
    try:
        if os.path.isfile(filename):
            workbook = load_workbook(filename)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([
                "column-1", "column-2"
            ])
        
        if isinstance(data, (list, tuple)):
            sheet.append(data)
        else:
            raise ValueError("Data should be a list or tuple")

        # Save the workbook to the file
        workbook.save(filename)
        print(f"Data successfully saved to '{filename}'")
    except Exception as e:
        print(f"Error saving to XLSX: {e}")